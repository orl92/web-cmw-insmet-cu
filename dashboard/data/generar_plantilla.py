import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

# Crear un archivo Excel
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Pronostico"

# Definir los encabezados y las posiciones de las tablas
headers = {
    'Fecha': 'A1',
    'Pronostico': 'A3',
    'Zona': 'A4',
    'Temperatura': 'B4',
    'Tiempo': 'E4',
    'Viento (DD)': 'H4',
    'Viento (FF)': 'K4',
    'Mar': 'N4',
    'Pronostico Extendido': 'A10',
    'Datos Astronómicos': 'F10'
}

# Definir las subcolumnas para "Temperatura", "Tiempo", "Viento (DD)", "Viento (FF)" y "Mar"
subcolumns = ['Mañana', 'Tarde (Máx)', 'Noche']

# Definir el estilo de borde
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Agregar los encabezados y las tablas al archivo Excel
for header, position in headers.items():
    cell = worksheet[position]
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Combinar celdas para los encabezados
worksheet.merge_cells('A4:A5')
worksheet.merge_cells('B4:D4')
worksheet.merge_cells('E4:G4')
worksheet.merge_cells('H4:J4')
worksheet.merge_cells('K4:M4')
worksheet.merge_cells('N4:P4')

# Agregar las subcolumnas
for col, subcol_start in zip(['B', 'E', 'H', 'K', 'N'], [2, 5, 8, 11, 14]):
    for i, subcol in enumerate(subcolumns):
        cell = worksheet.cell(row=5, column=subcol_start + i)
        cell.value = subcol
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# Agregar las filas de "Zona"
zones = ['Costa Norte', 'Interior', 'Costa Sur']
for i, zone in enumerate(zones, start=6):
    cell = worksheet.cell(row=i, column=1)
    cell.value = zone
    cell.border = thin_border

# Agregar bordes a las filas desde B6 hasta P8
for row in range(6, 9):
    for col in range(2, 17):
        cell = worksheet.cell(row=row, column=col)
        cell.border = thin_border

# Agregar filas adicionales para "Pronostico Extendido"
extended_columns = ['Día', 'Minima', 'Máxima', 'Tiempo']
for col_num, col_name in enumerate(extended_columns, start=1):
    cell = worksheet.cell(row=11, column=col_num)
    cell.value = col_name
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for i in range(1, 6):
    for j in range(1, 5):
        cell = worksheet.cell(row=11 + i, column=j)
        cell.value = ''
        cell.border = thin_border

# Agregar filas adicionales para "Datos Astronómicos"
astro_columns = ['Actual', 'Próxima', 'Fecha']
for col_num, col_name in enumerate(astro_columns, start=7):
    cell = worksheet.cell(row=11, column=col_num)
    cell.value = col_name
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

astro_data_rows = [
    ['Fase Lunar', '', '', ''],
    ['Salida Sol', '', '', ''],
    ['Puesta Sol', '', '', ''],
    ['Indice UV', '', '', '']
]

for i, row_data in enumerate(astro_data_rows, start=12):
    for j, value in enumerate(row_data, start=6):
        cell = worksheet.cell(row=i, column=j)
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# Ajustar el ancho de las columnas
for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column].width = adjusted_width

# Guardar el archivo Excel
workbook.save('plantilla_pronostico.xlsx')
